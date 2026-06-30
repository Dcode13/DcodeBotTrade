#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Scraper harga LBMA Gold (AM & PM) - default 6 bulan terakhir.

Sumber data: feed JSON publik resmi LBMA
  AM (auction 10:30 London): https://prices.lbma.org.uk/json/gold_am.json
  PM (auction 15:00 London): https://prices.lbma.org.uk/json/gold_pm.json

Setiap baris feed berbentuk:
  {"is_cms_locked": 0, "d": "YYYY-MM-DD", "v": [USD, GBP, EUR]}
Jadi v[0]=USD, v[1]=GBP, v[2]=EUR (EUR null untuk tanggal sebelum 1999).

Output: file Excel (.xlsx) dengan kolom:
  Date | AM_USD | AM_GBP | AM_EUR | PM_USD | PM_GBP | PM_EUR

Cara pakai:
  pip install requests pandas openpyxl
  python lbma_gold_scraper.py

Opsi:
  python lbma_gold_scraper.py --months 6 --output lbma_gold_6bulan.xlsx --csv
  python lbma_gold_scraper.py --start 2025-12-22 --end 2026-06-22

Catatan:
  - Harga LBMA Gold/Silver diadministrasikan oleh ICE Benchmark Administration (IBA).
    Feed di atas bebas diakses untuk keperluan pribadi/edukasi; untuk penggunaan
    komersial/redistribusi diperlukan lisensi dari IBA. Periksa ketentuan di situs LBMA/IBA.
"""

import argparse
import datetime as dt
import sys

import requests
import pandas as pd

AM_URL = "https://prices.lbma.org.uk/json/gold_am.json"
PM_URL = "https://prices.lbma.org.uk/json/gold_pm.json"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; lbma-gold-scraper/1.0; +https://www.lbma.org.uk/)",
    "Accept": "application/json, text/plain, */*",
}


def fetch_series(url, timeout=60):
    """Unduh satu feed JSON LBMA dan kembalikan list of dict."""
    resp = requests.get(url, headers=HEADERS, timeout=timeout)
    resp.raise_for_status()
    return resp.json()


def months_ago(d, months):
    """Tanggal `months` bulan sebelum d (clamp ke hari terakhir bila perlu)."""
    month_index = d.year * 12 + (d.month - 1) - months
    y, m = divmod(month_index, 12)
    m += 1
    if m == 12:
        first_next = dt.date(y + 1, 1, 1)
    else:
        first_next = dt.date(y, m + 1, 1)
    last_day = (first_next - dt.timedelta(days=1)).day
    return dt.date(y, m, min(d.day, last_day))


def _index_by_date(rows):
    """Map {date(): [USD, GBP, EUR]} dari list feed."""
    out = {}
    for row in rows:
        ds = row.get("d")
        if not ds:
            continue
        try:
            d = dt.date.fromisoformat(ds)
        except ValueError:
            continue
        v = row.get("v") or []
        v = (list(v) + [None, None, None])[:3]
        out[d] = v
    return out


def build_dataframe(am_rows, pm_rows, start, end):
    """Gabungkan AM & PM per tanggal, difilter [start, end] inklusif."""
    am = _index_by_date(am_rows)
    pm = _index_by_date(pm_rows)
    dates = sorted(d for d in (set(am) | set(pm)) if start <= d <= end)
    records = []
    for d in dates:
        a = am.get(d, [None, None, None])
        p = pm.get(d, [None, None, None])
        records.append({
            "Date": d.isoformat(),
            "AM_USD": a[0], "AM_GBP": a[1], "AM_EUR": a[2],
            "PM_USD": p[0], "PM_GBP": p[1], "PM_EUR": p[2],
        })
    cols = ["Date", "AM_USD", "AM_GBP", "AM_EUR", "PM_USD", "PM_GBP", "PM_EUR"]
    return pd.DataFrame.from_records(records, columns=cols)


def write_excel(df, path):
    """Tulis DataFrame ke .xlsx dengan format header & angka yang rapi."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        sheet = "LBMA Gold AM-PM"
        df.to_excel(xw, index=False, sheet_name=sheet)
        ws = xw.sheets[sheet]

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for col_idx, _ in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 13

        # angka 2 desimal untuk kolom harga (kolom 2..7)
        for col_idx in range(2, 8):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

        ws.freeze_panes = "B2"
    return path


def parse_args():
    ap = argparse.ArgumentParser(
        description="Scrape harga LBMA Gold AM & PM (default 6 bulan terakhir)."
    )
    ap.add_argument("--months", type=int, default=6,
                    help="Jumlah bulan ke belakang dari hari ini (default 6).")
    ap.add_argument("--start", default=None, help="Tanggal mulai YYYY-MM-DD (override --months).")
    ap.add_argument("--end", default=None, help="Tanggal akhir YYYY-MM-DD (default hari ini).")
    ap.add_argument("--output", default="lbma_gold_am_pm.xlsx", help="Nama file Excel keluaran.")
    ap.add_argument("--csv", action="store_true", help="Tulis juga file .csv.")
    return ap.parse_args()


def main():
    args = parse_args()
    end = dt.date.fromisoformat(args.end) if args.end else dt.date.today()
    start = dt.date.fromisoformat(args.start) if args.start else months_ago(end, args.months)

    print(f"Rentang: {start.isoformat()} s/d {end.isoformat()}")
    print("Mengunduh AM ...")
    am_rows = fetch_series(AM_URL)
    print("Mengunduh PM ...")
    pm_rows = fetch_series(PM_URL)

    df = build_dataframe(am_rows, pm_rows, start, end)
    if df.empty:
        print("Tidak ada data pada rentang tersebut.", file=sys.stderr)
        sys.exit(1)

    write_excel(df, args.output)
    print(f"OK: {len(df)} baris ({df['Date'].iloc[0]} -> {df['Date'].iloc[-1]}) disimpan ke {args.output}")

    if args.csv:
        csv_path = args.output.rsplit(".", 1)[0] + ".csv"
        df.to_csv(csv_path, index=False)
        print(f"CSV: {csv_path}")


if __name__ == "__main__":
    main()
